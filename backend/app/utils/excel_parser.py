"""Excel 测试用例解析器 — 智能列名匹配 + 3 种步骤格式解析"""

import io
import json
import re
import logging
from dataclasses import dataclass, field
from typing import Any, Optional

logger = logging.getLogger("autopilot.excel")

# ── 标准列名 → 别名映射 ──
COLUMN_ALIASES: dict[str, list[str]] = {
    "case_no":       ["用例编号", "编号", "caseno", "no", "id", "序号", "case id"],
    "case_name":     ["用例名称", "名称", "casename", "用例名", "title", "测试点", "测试用例", "用例标题", "标题", "case", "测试用例名称"],
    "priority":      ["优先级", "priority", "重要程度", "level", "级别", "等级", "用例级别"],
    "pre_condition": ["前置条件", "precondition", "预置条件", "前置", "前提条件", "准备", "前提"],
    "steps":         ["操作步骤", "steps", "测试步骤", "step", "步骤描述", "步骤", "测试操作", "操作步骤描述"],
    "expected_result": ["预期结果", "expectedresult", "期望结果", "expected", "预期", "期望", "预期输出"],
}

# ── 可合并为步骤的三列 ──
ACTION_COLUMNS = ["操作", "动作", "action", "操作类型"]
TARGET_COLUMNS = ["对象", "目标", "target", "操作对象", "元素", "定位"]
VALUE_COLUMNS  = ["数据", "值", "value", "输入值", "输入", "参数"]

# ── 11 种标准 action 枚举（Web + Android 通用） ──
VALID_ACTIONS = frozenset({
    "navigate", "fill", "click", "select", "hover",
    "assert_text", "assert_visible", "screenshot", "wait",
    "swipe", "back",
})

# ── 纯文本步骤模式 ──
LINE_PATTERN = re.compile(
    r'^(?:步骤\s*)?(\d+)[\.\、\s]\s*(.+)$',
    re.MULTILINE,
)

logger = logging.getLogger("autopilot.excel")


@dataclass
class ParsedStep:
    """解析后的单条步骤"""
    action: str
    target: str = ""
    value: str = ""
    description: str = ""


@dataclass
class ParsedCase:
    """解析后的一条用例"""
    case_name: str
    case_no: Optional[str] = None
    priority: str = "P1"
    pre_condition: Optional[str] = None
    steps: list[ParsedStep] = field(default_factory=list)
    expected_result: Optional[str] = None
    row_number: int = 0


@dataclass
class ParseResult:
    """一次解析的完整结果"""
    total_rows: int = 0
    success: int = 0
    failed: int = 0
    cases: list[ParsedCase] = field(default_factory=list)
    errors: list[dict] = field(default_factory=list)


class ExcelParser:
    """Excel 测试用例解析器"""

    @staticmethod
    def parse(file_content: bytes, filename: str = "") -> ParseResult:
        """主入口：解析 Excel 内容为用例列表

        Args:
            file_content: Excel 文件原始字节
            filename: 文件名（用于判断 .xls / .xlsx）

        Returns:
            ParseResult with parsed cases and error details
        """
        ws = _read_sheet(file_content, filename)
        if not ws:
            result = ParseResult()
            result.failed = 1
            result.errors.append({"row": 0, "reason": "无法读取工作表"})
            return result

        # 读 header
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            result = ParseResult()
            result.failed = 1
            result.errors.append({"row": 0, "reason": "Excel 至少需要 1 行表头 + 1 行数据"})
            return result

        headers = [str(h).strip() if h else "" for h in rows[0]]
        col_map = _match_columns(headers)

        # 若存在 action+target 合并列，且 steps 可能被误匹配，优先使用合并路径
        if "steps" in col_map and _can_merge_steps(headers):
            steps_col = headers[col_map["steps"]]
            # 检查 steps 列名是否实际上是 action/target/value 列
            if _is_merge_column(steps_col):
                del col_map["steps"]

        # 检查必填列
        missing = []
        for required in ["case_name", "steps"]:
            if required not in col_map:
                # 检查是否可通过三列合并
                if required == "steps" and _can_merge_steps(headers):
                    continue
                missing.append(required)
        if missing:
            result = ParseResult()
            result.failed = 1
            result.errors.append({
                "row": 0,
                "reason": f"缺少必填列: {', '.join(missing)}，请检查表头。已识别表头: {headers[:10]}",
            })
            return result

        merge_idx = _find_merge_indices(headers) if "steps" not in col_map else (None, None, None)
        action_idx, target_idx, value_idx = merge_idx

        result = ParseResult()

        for i, row in enumerate(rows[1:], start=2):
            result.total_rows += 1
            values = [str(v).strip() if v is not None else "" for v in row]

            try:
                case_name = _cell(row, col_map.get("case_name"))
                if not case_name:
                    result.failed += 1
                    result.errors.append({"row": i, "reason": "用例名称为空"})
                    continue

                # 解析步骤
                if "steps" in col_map:
                    steps_raw = _cell(row, col_map["steps"])
                    steps = _parse_steps(steps_raw)
                elif action_idx is not None and target_idx is not None:
                    steps_raw_action = _cell(row, action_idx) if action_idx < len(values) else ""
                    steps_raw_target = _cell(row, target_idx) if target_idx < len(values) else ""
                    steps_raw_value = _cell(row, value_idx) if value_idx is not None and value_idx < len(values) else ""
                    steps = _parse_steps_from_columns(steps_raw_action, steps_raw_target, steps_raw_value)
                else:
                    result.failed += 1
                    result.errors.append({"row": i, "reason": "无法找到步骤列"})
                    continue

                # 校验 steps
                valid_steps, step_errors = _validate_steps(steps, i)
                if step_errors:
                    result.failed += 1
                    for e in step_errors:
                        result.errors.append(e)
                    continue

                case = ParsedCase(
                    case_name=case_name,
                    case_no=_cell(row, col_map.get("case_no")),
                    priority=_normalize_priority(_cell(row, col_map.get("priority"))),
                    pre_condition=_cell(row, col_map.get("pre_condition")) or None,
                    steps=valid_steps,
                    expected_result=_cell(row, col_map.get("expected_result")) or None,
                    row_number=i,
                )
                result.cases.append(case)
                result.success += 1

            except Exception as e:
                result.failed += 1
                result.errors.append({"row": i, "reason": str(e)})

        return result


# ═══════════════════════════════════════════════
# 列名匹配
# ═══════════════════════════════════════════════

def _normalize_header(h: str) -> str:
    """标准化表头：去空格、去特殊字符、小写"""
    return re.sub(r'[\s_\-（）()【】\[\]{}]', '', h).lower()


def _match_columns(headers: list[str]) -> dict[str, int]:
    """智能匹配表头 → 标准列名映射（先精确，再模糊）

    合并列（操作/动作/对象/目标/数据/值）不参与标准列名匹配，
    以避免"操作"被误匹配为 case_no 或 steps。
    """
    result: dict[str, int] = {}
    used_indices: set[int] = set()
    all_norms = [_normalize_header(h) for h in headers]

    # 标记合并列索引（这些列不参与标准匹配）
    merge_indices: set[int] = set()
    for idx, norm in enumerate(all_norms):
        if _is_merge_column(norm):
            merge_indices.add(idx)

    # Pass 1: 精确匹配
    for std_name, aliases in COLUMN_ALIASES.items():
        alias_norms = [_normalize_header(a) for a in aliases]
        for idx, norm in enumerate(all_norms):
            if idx in used_indices or idx in merge_indices:
                continue
            if norm in alias_norms:
                result[std_name] = idx
                used_indices.add(idx)
                break

    # Pass 2: 模糊匹配（剩余列、剩余索引）
    for std_name, aliases in COLUMN_ALIASES.items():
        if std_name in result:
            continue
        alias_norms = [_normalize_header(a) for a in aliases]
        for idx, norm in enumerate(all_norms):
            if idx in used_indices or idx in merge_indices:
                continue
            if any(
                _levenshtein(norm, an) <= 2 or an in norm or norm in an
                for an in alias_norms
            ):
                result[std_name] = idx
                used_indices.add(idx)
                break

    return result


def _is_merge_column(col_name: str) -> bool:
    """判断列名是否属于 action/target/value 合并列"""
    norm = _normalize_header(col_name)
    all_merge = ACTION_COLUMNS + TARGET_COLUMNS + VALUE_COLUMNS
    return any(_normalize_header(a) == norm or _levenshtein(norm, _normalize_header(a)) <= 1 for a in all_merge)


def _can_merge_steps(headers: list[str]) -> bool:
    """判断是否可通过 action + target + value 三列合并"""
    norms = [_normalize_header(h) for h in headers]

    def _match_any(col_name: str, candidates: list[str]) -> bool:
        cn = _normalize_header(col_name)
        for c in candidates:
            cn2 = _normalize_header(c)
            if cn == cn2 or _levenshtein(cn, cn2) <= 1:
                return True
            if len(cn) >= 2 and (cn.startswith(cn2) or cn2.startswith(cn)):
                return True
        return False

    has_action = any(_match_any(n, ACTION_COLUMNS) for n in norms)
    has_target = any(_match_any(n, TARGET_COLUMNS) for n in norms)
    return has_action and has_target


def _find_merge_indices(headers: list[str]) -> tuple[int | None, int | None, int | None]:
    """定位 action/target/value 三列索引（需要精确匹配，避免子串混淆）"""
    norms = [_normalize_header(h) for h in headers]

    def _match(col_name: str, candidates: list[str]) -> bool:
        cn = _normalize_header(col_name)
        for c in candidates:
            cn2 = _normalize_header(c)
            if cn == cn2:
                return True
            if _levenshtein(cn, cn2) <= 1:
                return True
        return False

    action_idx = next((i for i, n in enumerate(norms) if _match(n, ACTION_COLUMNS)), None)
    target_idx = next((i for i, n in enumerate(norms) if _match(n, TARGET_COLUMNS)), None)
    value_idx = next((i for i, n in enumerate(norms) if _match(n, VALUE_COLUMNS)), None)

    return action_idx, target_idx, value_idx


# ═══════════════════════════════════════════════
# 步骤解析（3 种格式）
# ═══════════════════════════════════════════════

def _parse_steps(raw: str) -> list[ParsedStep]:
    """解析步骤，自动判断格式"""
    if not raw:
        return []

    # 格式一：JSON 数组
    raw_stripped = raw.strip()
    if raw_stripped.startswith("[") and raw_stripped.endswith("]"):
        try:
            data = json.loads(raw_stripped)
            if isinstance(data, list) and all(isinstance(d, dict) for d in data):
                return [ParsedStep(
                    action=d.get("action", "action"),
                    target=d.get("target", ""),
                    value=d.get("value", ""),
                    description=d.get("description", d.get("desc", "")),
                ) for d in data]
        except json.JSONDecodeError:
            pass

    # 格式二：纯文本多行
    lines = [l.strip() for l in raw.split("\n") if l.strip()]
    if len(lines) > 0 and lines[0][0].isdigit():
        return _parse_text_lines(lines)

    # 兜底：每行作为一个步骤
    return [ParsedStep(action="action", target="", value=l, description=l) for l in lines if l]


def _parse_text_lines(lines: list[str]) -> list[ParsedStep]:
    """解析纯文本步骤行

    示例:
        1. 打开登录页面 https://example.com/login
        2. 在用户名输入框输入 admin
        3. 点击登录按钮
        4. 验证页面显示"欢迎"
    """
    steps: list[ParsedStep] = []
    for line in lines:
        if not line:
            continue
        action, target, value, desc = _classify_text_step(line)
        steps.append(ParsedStep(action=action, target=target, value=value, description=desc))
    return steps


def _classify_text_step(line: str) -> tuple[str, str, str, str]:
    """基于关键词分类单条文本步骤"""
    clean = re.sub(r'^\d+[\.\、\s]+', '', line).strip()
    desc = clean[:200]

    # navigate: 打开/访问/跳转 + URL
    if re.search(r'打开|访问|跳转|进入|navigate|goto|浏览', clean):
        url_match = re.search(r'(https?://\S+)', clean)
        if url_match:
            return ("navigate", url_match.group(1), "", desc)
        return ("navigate", clean, "", desc)

    # fill: 输入/填写 + 值
    if re.search(r'输入|填写|填入|键入|fill|input|enter', clean):
        parts = re.split(r'输入|填写|填入|键入', clean, maxsplit=1)
        target = parts[0].replace("在", "").replace("向", "").strip() if len(parts) > 0 else ""
        value = parts[1].strip() if len(parts) > 1 else ""
        return ("fill", target, value, desc)

    # click: 点击/按下/选择按钮
    if re.search(r'点击|按下|click|press|tap|选择.*按钮', clean):
        target = re.sub(r'点击|按下|click|press|tap', '', clean).strip()
        return ("click", target, "", desc)

    # select: 下拉选择
    if re.search(r'下拉|选择|select|picker', clean):
        target = re.sub(r'下拉|选择|select|picker', '', clean).strip()
        return ("select", target, "", desc)

    # hover: 悬停
    if re.search(r'悬停|hover|鼠标移', clean):
        return ("hover", clean, "", desc)

    # assert: 验证/断言/检查/确认
    if re.search(r'验证|断言|检查|确认|assert|expect|verify|check|应该', clean):
        return ("assert_text", clean, "", desc)

    # screenshot: 截图
    if re.search(r'截图|screenshot|capture', clean, re.IGNORECASE):
        return ("screenshot", clean, "", desc)

    # wait: 等待
    if re.search(r'等待|sleep|wait|延时', clean):
        return ("wait", clean, "", desc)

    # 兜底
    return ("action", clean, "", desc)


def _parse_steps_from_columns(action_col: str, target_col: str, value_col: str) -> list[ParsedStep]:
    """将 action + target + value 三列合并为步骤"""
    action_lines = [l.strip() for l in action_col.split("\n") if l.strip()] if action_col else []
    target_lines = [l.strip() for l in target_col.split("\n") if l.strip()] if target_col else []
    value_lines = [l.strip() for l in value_col.split("\n") if l.strip()] if value_col else []

    # 如果只有一行，直接合并
    if len(action_lines) <= 1 and len(target_lines) <= 1 and len(value_lines) <= 1:
        return [ParsedStep(
            action=_normalize_action(action_col),
            target=target_col,
            value=value_col,
            description=f"{action_col} {target_col} {value_col}".strip(),
        )]

    # 多行 → 按行对应
    max_lines = max(len(action_lines), len(target_lines), len(value_lines))
    steps: list[ParsedStep] = []
    for i in range(max_lines):
        a = action_lines[i] if i < len(action_lines) else ""
        t = target_lines[i] if i < len(target_lines) else ""
        v = value_lines[i] if i < len(value_lines) else ""
        steps.append(ParsedStep(
            action=_normalize_action(a),
            target=t,
            value=v,
            description=f"{a} {t} {v}".strip(),
        ))
    return steps


# ═══════════════════════════════════════════════
# 校验
# ═══════════════════════════════════════════════

def _validate_steps(steps: list[ParsedStep], row_num: int) -> tuple[list[ParsedStep], list[dict]]:
    """校验步骤：action 枚举检查，自动补全"""
    errors = []
    validated = []

    for j, s in enumerate(steps):
        action = _normalize_action(s.action)
        if action not in VALID_ACTIONS:
            errors.append({"row": row_num, "reason": f"步骤 {j+1}: action '{s.action}' 不是有效枚举值，应为: {', '.join(sorted(VALID_ACTIONS))}"})
            continue
        validated.append(ParsedStep(
            action=action,
            target=s.target,
            value=s.value,
            description=s.description or f"{action} {s.target}".strip(),
        ))

    if not validated and not errors:
        errors.append({"row": row_num, "reason": "步骤列表为空"})

    return validated, errors


def _normalize_action(action: str) -> str:
    """标准化 action 名称（中文 → 英文枚举）"""
    a = action.strip().lower()
    mapping = {
        "navigate": "navigate", "导航": "navigate", "打开": "navigate", "访问": "navigate", "跳转": "navigate",
        "fill": "fill", "输入": "fill", "填写": "fill", "键入": "fill",
        "click": "click", "点击": "click", "按下": "click",
        "select": "select", "选择": "select", "下拉": "select",
        "hover": "hover", "悬停": "hover",
        "assert_text": "assert_text", "断言": "assert_text", "验证": "assert_text", "检查": "assert_text", "assert": "assert_text",
        "assert_visible": "assert_visible", "可见": "assert_visible", "显示": "assert_visible",
        "screenshot": "screenshot", "截图": "screenshot",
        "wait": "wait", "等待": "wait", "延时": "wait", "暂停": "wait",
        "swipe": "swipe", "滑动": "swipe", "滑": "swipe",
        "back": "back", "返回": "back", "后退": "back",
    }
    return mapping.get(a, a)


def _normalize_priority(raw: str | None) -> str:
    """标准化优先级"""
    if not raw:
        return "P1"
    p = raw.strip().upper()
    if p in ("P0", "P1", "P2", "P3"):
        return p
    # 数字映射
    num_map = {"0": "P0", "1": "P1", "2": "P2", "3": "P3", "高": "P0", "中": "P2", "低": "P3", "HIGH": "P0", "MEDIUM": "P2", "LOW": "P3"}
    return num_map.get(p, "P1")


# ═══════════════════════════════════════════════
# 工具函数
# ═══════════════════════════════════════════════

def _read_sheet(file_content: bytes, filename: str = ""):
    """读取工作表（支持 .xlsx 和 .xls）"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_content), read_only=True)
        return wb.active
    except Exception:
        # 尝试 xlrd for .xls
        if filename.lower().endswith(".xls"):
            try:
                import xlrd
                wb = xlrd.open_workbook(file_contents=file_content)
                return _xlrd_to_rows(wb.sheet_by_index(0))
            except ImportError:
                raise RuntimeError("解析 .xls 需要安装 xlrd: pip install xlrd")
            except Exception:
                pass
        raise


def _xlrd_to_rows(sheet) -> list:
    """将 xlrd sheet 转为 openpyxl 兼容的 rows 格式"""
    rows = []
    for r in range(sheet.nrows):
        rows.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])
    # 包装成类似 openpyxl 的 iter_rows 行为
    class _FakeWS:
        @staticmethod
        def iter_rows(values_only=False):
            return rows
    return _FakeWS()


def _cell(row, col_idx: int | None) -> str:
    """安全取单元格值"""
    if col_idx is None or col_idx >= len(row):
        return ""
    v = row[col_idx]
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


# ═══════════════════════════════════════════════
# Levenshtein 距离
# ═══════════════════════════════════════════════

def _levenshtein(s1: str, s2: str) -> int:
    """计算两个字符串的编辑距离"""
    if len(s1) < len(s2):
        return _levenshtein(s2, s1)
    if len(s2) == 0:
        return len(s1)

    prev = list(range(len(s2) + 1))
    for i, c1 in enumerate(s1):
        curr = [i + 1]
        for j, c2 in enumerate(s2):
            curr.append(min(
                prev[j + 1] + 1,      # insert
                curr[j] + 1,          # delete
                prev[j] + (0 if c1 == c2 else 1),  # substitute
            ))
        prev = curr
    return prev[-1]
