"""ExcelParser unit tests — Excel test case parsing."""

import io
import json
import pytest
from openpyxl import Workbook

from app.utils.excel_parser import ExcelParser, ParseResult, ParsedCase, ParsedStep


def _make_excel(headers: list[str], rows: list[list]) -> io.BytesIO:
    """Create an in-memory .xlsx workbook and return a BytesIO buffer."""
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class TestParseJsonSteps:
    """Parse JSON steps format (steps column contains JSON array)"""

    def test_json_steps_parsed(self):
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["测试用例1", json.dumps([
                    {"action": "click", "target": "#btn", "value": ""},
                    {"action": "fill", "target": "#input", "value": "hello"},
                ]), "P0"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        assert result.failed == 0
        assert len(result.cases) == 1
        case = result.cases[0]
        assert case.case_name == "测试用例1"
        assert len(case.steps) == 2
        assert case.steps[0].action == "click"
        assert case.steps[0].target == "#btn"
        assert case.steps[1].action == "fill"
        assert case.steps[1].value == "hello"

    def test_json_steps_with_description(self):
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["测试用例", json.dumps([
                    {"action": "navigate", "target": "https://example.com", "description": "打开首页"},
                ]), "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        assert result.cases[0].steps[0].description == "打开首页"


class TestParsePlainTextSteps:
    """Parse plain text steps (line-based numbered format)"""

    def test_numbered_text_steps(self):
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["测试用例", "1. 打开 https://example.com\n2. 点击登录按钮\n3. 输入用户名 admin", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        assert len(result.cases[0].steps) == 3

    def test_numbered_steps_with_chinese_dot(self):
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["测试用例", "1、打开百度\n2、点击搜索按钮", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        assert len(result.cases[0].steps) == 2


class TestParseActionTargetValueColumns:
    """Parse action+target+value columns (3 columns merged)"""

    def test_three_column_merge(self):
        buf = _make_excel(
            ["用例名称", "操作", "对象", "优先级"],
            [
                ["测试用例", "点击", "#btn", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        assert len(result.cases[0].steps) == 1
        assert result.cases[0].steps[0].action == "click"

    def test_three_column_with_value(self):
        buf = _make_excel(
            ["用例名称", "操作", "对象", "数据", "优先级"],
            [
                ["测试用例", "输入", "#username", "admin", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        step = result.cases[0].steps[0]
        assert step.action == "fill"
        assert step.target == "#username"
        assert step.value == "admin"

    def test_three_column_english_headers(self):
        buf = _make_excel(
            ["用例名称", "action", "target", "value", "优先级"],
            [
                ["Test Case", "click", "#btn", "", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        assert result.cases[0].steps[0].action == "click"


class TestChineseColumnNames:
    """Chinese column name recognition"""

    def test_chinese_column_names(self):
        buf = _make_excel(
            ["用例编号", "用例名称", "操作步骤", "优先级", "前置条件", "预期结果"],
            [
                ["TC001", "登录测试", json.dumps([{"action": "navigate", "target": "/login"}]), "P0", "已注册账号", "登录成功"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        case = result.cases[0]
        assert case.case_no == "TC001"
        assert case.case_name == "登录测试"
        assert case.priority == "P0"
        assert case.pre_condition == "已注册账号"
        assert case.expected_result == "登录成功"

    def test_alternative_chinese_column_names(self):
        buf = _make_excel(
            ["编号", "测试用例名称", "测试步骤", "级别", "预期"],
            [
                ["TC002", "注册测试", json.dumps([{"action": "click", "target": "#register"}]), "P2", "注册成功"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        case = result.cases[0]
        assert case.case_no == "TC002"
        assert case.case_name == "注册测试"
        assert case.priority == "P2"


class TestFuzzyColumnMatching:
    """Fuzzy column matching via Levenshtein distance"""

    def test_slight_typo_in_header(self):
        buf = _make_excel(
            ["用例名称", "操做步骤", "优先级"],  # 操作 → 操做 (typo)
            [
                ["测试用例", json.dumps([{"action": "click", "target": "#btn"}]), "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1

    def test_extra_spaces_in_header(self):
        buf = _make_excel(
            ["用例名称", "操作步骤 ", "优先级"],
            [
                ["测试用例", json.dumps([{"action": "click", "target": "#btn"}]), "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1


class TestActionNormalization:
    """Action normalization (Chinese → English)"""

    @pytest.mark.parametrize("chinese,expected", [
        ("点击", "click"),
        ("输入", "fill"),
        ("打开", "navigate"),
        ("导航", "navigate"),
        ("选择", "select"),
        ("悬停", "hover"),
        ("验证", "assert_text"),
        ("截图", "screenshot"),
        ("等待", "wait"),
    ])
    def test_chinese_action_normalized(self, chinese, expected):
        buf = _make_excel(
            ["用例名称", "操作", "对象", "优先级"],
            [
                ["测试", chinese, "#el", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        assert result.cases[0].steps[0].action == expected

    def test_english_action_preserved(self):
        buf = _make_excel(
            ["用例名称", "操作", "对象", "优先级"],
            [
                ["测试", "navigate", "https://example.com", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        assert result.cases[0].steps[0].action == "navigate"


class TestPriorityNormalization:
    """Priority normalization"""

    @pytest.mark.parametrize("raw,expected", [
        ("高", "P0"),
        ("HIGH", "P0"),
        ("中", "P2"),
        ("低", "P3"),
        ("P0", "P0"),
        ("P1", "P1"),
        ("P2", "P2"),
        ("P3", "P3"),
        ("", "P1"),
    ])
    def test_priority_normalized(self, raw, expected):
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["测试", json.dumps([{"action": "click", "target": "#btn"}]), raw],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        assert result.cases[0].priority == expected


class TestDuplicateCaseNo:
    """Duplicate case_no detection — parser accepts duplicates"""

    def test_duplicate_case_no_accepted(self):
        buf = _make_excel(
            ["用例编号", "用例名称", "操作步骤", "优先级"],
            [
                ["TC001", "用例A", json.dumps([{"action": "click", "target": "#a"}]), "P1"],
                ["TC001", "用例B", json.dumps([{"action": "click", "target": "#b"}]), "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.total_rows == 2
        assert result.success == 2
        assert result.cases[0].case_no == "TC001"
        assert result.cases[1].case_no == "TC001"


class TestEmptyFileValidation:
    """Empty file / insufficient rows validation"""

    def test_only_header_row_returns_error(self):
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [],  # no data rows
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.failed == 1
        assert len(result.errors) > 0

    def test_empty_case_name_counted_as_failed(self):
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["", json.dumps([{"action": "click", "target": "#btn"}]), "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.failed >= 1


class TestMissingRequiredColumns:
    """Missing required columns validation"""

    def test_missing_case_name_column(self):
        buf = _make_excel(
            ["未知列", "操作步骤", "优先级"],
            [
                ["数据", json.dumps([{"action": "click", "target": "#btn"}]), "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.failed == 1
        assert "必填列" in result.errors[0]["reason"]

    def test_missing_steps_column_without_merge(self):
        buf = _make_excel(
            ["用例名称", "未知列", "优先级"],
            [
                ["测试", "数据", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.failed == 1


class TestNonXlsxValidation:
    """Non-xlsx / invalid file handling"""

    def test_invalid_bytes_raises_exception(self):
        with pytest.raises(Exception):
            ExcelParser.parse(b"not an excel file at all", "test.xlsx")

    def test_garbage_bytes_raises(self):
        with pytest.raises(Exception):
            ExcelParser.parse(b"\x00\x01\x02\x03", "test.xlsx")


class TestMultipleRows:
    """Multiple data rows parsing"""

    def test_multiple_cases_parsed(self):
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["用例A", json.dumps([{"action": "click", "target": "#a"}]), "P0"],
                ["用例B", json.dumps([{"action": "fill", "target": "#b", "value": "x"}]), "P1"],
                ["用例C", json.dumps([{"action": "navigate", "target": "/"}]), "P2"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.total_rows == 3
        assert result.success == 3
        assert len(result.cases) == 3

    def test_mixed_success_and_failure(self):
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["用例A", json.dumps([{"action": "click", "target": "#a"}]), "P1"],
                ["", json.dumps([{"action": "click", "target": "#b"}]), "P1"],  # empty name
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.total_rows == 2
        assert result.success == 1
        assert result.failed >= 1


class TestXlsExtension:
    """XLS extension handling"""

    def test_xls_extension_tries_xlrd(self):
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["测试", json.dumps([{"action": "click", "target": "#btn"}]), "P1"],
            ],
        )
        # .xlsx content with .xls extension — openpyxl should still handle it
        result = ExcelParser.parse(buf.read(), "test.xls")
        assert result.success == 1


# ═══════════════════════════════════════════════
# 新增测试 — 覆盖未覆盖行
# ═══════════════════════════════════════════════


class TestReadSheetReturnsNone:
    """Cover lines 89-92: _read_sheet returns None → 无法读取工作表"""

    def test_read_sheet_returns_none(self, mocker):
        mocker.patch("app.utils.excel_parser._read_sheet", return_value=None)
        result = ExcelParser.parse(b"any content", "test.xlsx")
        assert result.failed == 1
        assert "无法读取工作表" in result.errors[0]["reason"]


class TestStepsColumnIsMergeColumn:
    """Cover lines 107-110: steps column is actually a merge column (动作/操作)"""

    def test_steps_column_is_merge_column(self):
        # "操作" is both in steps aliases (via fuzzy "操作步骤" contains "操作")
        # AND is a merge column (action column). The parser should detect this
        # and prefer the merge-column path.
        buf = _make_excel(
            ["用例名称", "操作", "对象", "优先级"],
            [
                ["测试用例", "点击", "#btn", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        assert result.cases[0].steps[0].action == "click"


class TestNoStepsAndNoMerge:
    """Cover lines 155-157: no steps column and no merge columns → 无法找到步骤列"""

    def test_no_steps_no_merge_columns(self):
        # _can_merge_steps returns True (action/target headers exist via startswith),
        # but _find_merge_indices returns None (Levenshtein > 1, no exact match),
        # so the code reaches the "无法找到步骤列" branch.
        buf = _make_excel(
            ["用例名称", "动作类型", "目标类型", "优先级"],
            [
                ["测试用例", "点击", "#btn", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.failed == 1
        assert "无法找到步骤列" in result.errors[0]["reason"]


class TestStepValidationErrors:
    """Cover lines 162-165, 431-432: step validation errors (invalid action)"""

    def test_invalid_action_in_json_steps(self):
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["测试用例", json.dumps([{"action": "invalid_action", "target": "#btn"}]), "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.failed == 1
        assert any("不是有效枚举值" in e["reason"] for e in result.errors)

    def test_invalid_action_in_merge_columns(self):
        buf = _make_excel(
            ["用例名称", "操作", "对象", "优先级"],
            [
                ["测试用例", "未知操作", "#btn", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.failed == 1
        assert any("不是有效枚举值" in e["reason"] for e in result.errors)


class TestRowProcessingException:
    """Cover lines 179-181: exception during row processing"""

    def test_exception_during_row_processing(self, mocker):
        # Mock _parse_steps to raise an exception for a specific row
        original = __import__("app.utils.excel_parser", fromlist=["_parse_steps"])._parse_steps
        call_count = [0]

        def mock_parse_steps(raw):
            call_count[0] += 1
            if call_count[0] == 1:
                return original(raw)
            raise RuntimeError("模拟的解析错误")

        mocker.patch("app.utils.excel_parser._parse_steps", side_effect=mock_parse_steps)

        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["用例A", json.dumps([{"action": "click", "target": "#a"}]), "P1"],
                ["用例B", json.dumps([{"action": "click", "target": "#b"}]), "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        # First row should succeed, second should fail
        assert result.success >= 1
        assert result.failed >= 1
        assert any("模拟的解析错误" in e["reason"] for e in result.errors)


class TestEmptySteps:
    """Cover line 295: empty steps raw → returns []"""

    def test_empty_steps_cell(self):
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["测试用例", "", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        # Empty steps → _validate_steps returns empty → error at line 441
        assert result.failed == 1
        assert any("步骤列表为空" in e["reason"] for e in result.errors)


class TestMalformedJsonSteps:
    """Cover lines 309-310: JSON decode error → pass (fall through to text parsing)"""

    def test_malformed_json_falls_through(self):
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["测试用例", '[{"action": "click", "target": "#btn"', "P1"],  # missing closing bracket
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        # Malformed JSON → falls through to text parsing → each line as a step
        # Fallback action="action" is not in VALID_ACTIONS → validation fails
        assert result.failed == 1

    def test_json_array_with_invalid_syntax(self):
        """Cover 309-310: [ ... ] that is not valid JSON → JSONDecodeError → pass"""
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["测试用例", '[{"action": click, "target": "#btn"}]', "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        # Invalid JSON inside brackets → JSONDecodeError → falls through to text
        assert result.failed == 1

    def test_json_array_with_non_dict_items(self):
        """Cover 302->313: JSON array where not all items are dicts"""
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["测试用例", '[{"action": "click", "target": "#btn"}, "not_a_dict"]', "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        # Not all items are dicts → falls through to text parsing
        assert result.failed == 1


class TestFallbackPlainTextPerLine:
    """Cover line 318: fallback — each line as a step with action="action" """

    def test_plain_text_per_line_fallback(self):
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["测试用例", "执行操作A\n执行操作B", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        # Fallback: each non-numeric line becomes a step with action="action"
        # "action" is not in VALID_ACTIONS → validation fails
        assert result.failed == 1
        assert result.total_rows == 1


class TestTextStepsEmptyLine:
    """Cover line 333: empty line between numbered steps (continue)"""

    def test_empty_line_between_numbered_steps(self):
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["测试用例", "1. 打开百度\n\n2. 点击搜索按钮", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        assert len(result.cases[0].steps) == 2


class TestClassifyTextStepBranches:
    """Cover lines 364-385: various text step classification branches"""

    def test_select_action(self):
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["测试用例", "1. 下拉选择城市", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        assert result.cases[0].steps[0].action == "select"

    def test_hover_action(self):
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["测试用例", "1. 悬停到菜单上", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        assert result.cases[0].steps[0].action == "hover"

    def test_assert_text_action(self):
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["测试用例", "1. 验证页面显示欢迎信息", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        assert result.cases[0].steps[0].action == "assert_text"

    def test_screenshot_action(self):
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["测试用例", "1. 截图保存当前页面", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        assert result.cases[0].steps[0].action == "screenshot"

    def test_wait_action(self):
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["测试用例", "1. 等待3秒", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        assert result.cases[0].steps[0].action == "wait"

    def test_fallback_text_action(self):
        # A line that doesn't match any keyword → action="action"
        # "action" is not in VALID_ACTIONS → validation fails
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["测试用例", "1. 执行自定义脚本", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.failed == 1
        assert any("不是有效枚举值" in e["reason"] for e in result.errors)


class TestMultiLineMergeColumns:
    """Cover lines 404-416: multi-line action/target/value merge"""

    def test_multi_line_action_target_value(self):
        buf = _make_excel(
            ["用例名称", "操作", "对象", "数据", "优先级"],
            [
                ["测试用例", "点击\n输入", "#btn\n#username", "val1\nval2", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        assert len(result.cases[0].steps) == 2
        assert result.cases[0].steps[0].action == "click"
        assert result.cases[0].steps[0].target == "#btn"
        assert result.cases[0].steps[0].value == "val1"
        assert result.cases[0].steps[1].action == "fill"
        assert result.cases[0].steps[1].target == "#username"
        assert result.cases[0].steps[1].value == "val2"


class TestEmptyStepsAfterValidation:
    """Cover line 441: empty steps list after validation"""

    def test_empty_steps_after_validation(self):
        # Steps cell contains only whitespace → _parse_steps returns [] → validated is empty
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["测试用例", "   ", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.failed == 1
        assert any("步骤列表为空" in e["reason"] for e in result.errors)


class TestXlrdFallback:
    """Cover lines 488-495: xlrd import path for .xls files"""

    def test_xls_with_xlrd_not_installed(self, mocker):
        # Mock openpyxl.load_workbook to fail (imported inside _read_sheet)
        mocker.patch("openpyxl.load_workbook", side_effect=Exception("openpyxl failed"))

        # Make xlrd import fail
        import builtins
        real_import = builtins.__import__

        def mock_import(name, *args, **kwargs):
            if name == "xlrd":
                raise ImportError("No module named xlrd")
            return real_import(name, *args, **kwargs)

        mocker.patch("builtins.__import__", side_effect=mock_import)

        with pytest.raises(RuntimeError, match="需要安装 xlrd"):
            ExcelParser.parse(b"dummy content", "test.xls")

    def test_xls_with_xlrd_available(self, mocker):
        """Cover lines 501-509: _xlrd_to_rows for xlrd sheet conversion"""
        mocker.patch("openpyxl.load_workbook", side_effect=Exception("openpyxl failed"))

        # Create a mock xlrd module and sheet
        mock_xlrd = mocker.MagicMock()
        mock_sheet = mocker.MagicMock()
        mock_sheet.nrows = 2
        mock_sheet.ncols = 3
        mock_sheet.cell_value.side_effect = lambda r, c: [
            ["用例名称", "操作步骤", "优先级"],
            ["测试", json.dumps([{"action": "click", "target": "#btn"}]), "P1"],
        ][r][c]
        mock_xlrd.open_workbook.return_value.sheet_by_index.return_value = mock_sheet

        mocker.patch.dict("sys.modules", {"xlrd": mock_xlrd})

        result = ExcelParser.parse(b"dummy content", "test.xls")
        assert result.success == 1
        assert result.cases[0].case_name == "测试"


class TestFloatCellValue:
    """Cover line 520: float cell value conversion (e.g., 1.0 → "1")"""

    def test_float_cell_value_via_cell_function(self):
        from app.utils.excel_parser import _cell
        # Directly test _cell with a float that equals its int value
        assert _cell([1.0], 0) == "1"
        assert _cell([2.0], 0) == "2"
        # Non-integer float should use default str()
        assert _cell([1.5], 0) == "1.5"

    def test_float_cell_value_in_excel_parsing(self):
        buf = _make_excel(
            ["用例编号", "用例名称", "操作步骤", "优先级"],
            [
                [1.0, "测试用例", json.dumps([{"action": "click", "target": "#btn"}]), "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1


class TestLevenshteinEmpty:
    """Cover line 533: Levenshtein distance base case (len(s2) == 0)"""

    def test_levenshtein_empty_s2(self):
        from app.utils.excel_parser import _levenshtein
        assert _levenshtein("abc", "") == 3
        assert _levenshtein("", "") == 0


class TestFindMergeIndicesLevenshtein:
    """Cover line 278: Levenshtein matching in _find_merge_indices"""

    def test_merge_indices_with_typo_in_action_header(self):
        # "操做" is a typo of "操作" — edit distance 1
        buf = _make_excel(
            ["用例名称", "操做", "对象", "优先级"],
            [
                ["测试用例", "点击", "#btn", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        assert result.cases[0].steps[0].action == "click"

    def test_merge_indices_with_typo_in_target_header(self):
        buf = _make_excel(
            ["用例名称", "操作", "对像", "优先级"],  # "对象" → "对像"
            [
                ["测试用例", "点击", "#btn", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        assert result.cases[0].steps[0].action == "click"


class TestNavigateWithoutUrl:
    """Cover text step navigate without URL (line 349)"""

    def test_navigate_without_url(self):
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["测试用例", "1. 打开登录页面", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        assert result.cases[0].steps[0].action == "navigate"


class TestFillTextStep:
    """Cover text step fill with 输入 keyword (lines 352-356)"""

    def test_fill_text_step_with_target_and_value(self):
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["测试用例", "1. 在搜索框输入 hello", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        step = result.cases[0].steps[0]
        assert step.action == "fill"
        assert "hello" in step.value


class TestAssertVisibleAction:
    """Cover assert_visible action normalization"""

    def test_assert_visible_action(self):
        buf = _make_excel(
            ["用例名称", "操作", "对象", "优先级"],
            [
                ["测试用例", "可见", "#element", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        assert result.cases[0].steps[0].action == "assert_visible"


class TestPriorityNumericMapping:
    """Cover numeric priority mapping (line 471)"""

    @pytest.mark.parametrize("raw,expected", [
        ("0", "P0"),
        ("1", "P1"),
        ("2", "P2"),
        ("3", "P3"),
        ("MEDIUM", "P2"),
        ("LOW", "P3"),
    ])
    def test_priority_numeric_and_english(self, raw, expected):
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["测试", json.dumps([{"action": "click", "target": "#btn"}]), raw],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        assert result.cases[0].priority == expected

    def test_priority_unknown_fallback(self):
        buf = _make_excel(
            ["用例名称", "操作步骤", "优先级"],
            [
                ["测试", json.dumps([{"action": "click", "target": "#btn"}]), "UNKNOWN"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        assert result.cases[0].priority == "P1"


class TestActionNormalizationExtra:
    """Cover additional action normalization mappings"""

    @pytest.mark.parametrize("raw,expected", [
        ("按下", "click"),
        ("填写", "fill"),
        ("键入", "fill"),
        ("断言", "assert_text"),
        ("assert", "assert_text"),
        ("显示", "assert_visible"),
        ("延时", "wait"),
        ("暂停", "wait"),
        ("访问", "navigate"),
        ("跳转", "navigate"),
    ])
    def test_extra_action_normalized(self, raw, expected):
        buf = _make_excel(
            ["用例名称", "操作", "对象", "优先级"],
            [
                ["测试", raw, "#el", "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1
        assert result.cases[0].steps[0].action == expected


class TestFuzzyMatchSubstring:
    """Cover fuzzy matching via substring (an in norm or norm in an)"""

    def test_header_contains_alias(self):
        # "测试用例名称" contains "用例名称" → should match case_name
        buf = _make_excel(
            ["测试用例名称", "操作步骤", "优先级"],
            [
                ["测试用例", json.dumps([{"action": "click", "target": "#btn"}]), "P1"],
            ],
        )
        result = ExcelParser.parse(buf.read(), "test.xlsx")
        assert result.success == 1


class TestXlsWithXlrdGeneralException:
    """Cover line 494-495: general exception in xlrd path (pass → re-raise)"""

    def test_xls_xlrd_general_exception(self, mocker):
        mocker.patch("openpyxl.load_workbook", side_effect=Exception("openpyxl failed"))

        mock_xlrd = mocker.MagicMock()
        mock_xlrd.open_workbook.side_effect = Exception("xlrd also failed")
        mocker.patch.dict("sys.modules", {"xlrd": mock_xlrd})

        with pytest.raises(Exception):
            ExcelParser.parse(b"dummy content", "test.xls")