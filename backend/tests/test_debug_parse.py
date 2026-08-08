"""Debug the parser directly"""
import io
from openpyxl import Workbook
from app.utils.excel_parser import _match_columns, _find_merge_indices, _cell, _is_merge_column

# Create merge-format Excel
wb = Workbook()
ws = wb.active
headers = ['用例名', '操作', '对象', '数据', '期望']
ws.append(headers)
ws.append(['注册新用户', 'fill', '用户名输入框', 'newuser', '注册成功提示'])
ws.append(['同意协议', 'click', '同意复选框', '', '协议已勾选'])
buf = io.BytesIO()
wb.save(buf)

from openpyxl import load_workbook
wb2 = load_workbook(io.BytesIO(buf.getvalue()), read_only=True)
ws2 = wb2.active
rows = list(ws2.iter_rows(values_only=True))

# Debug headers
headers_raw = [str(h).strip() if h else "" for h in rows[0]]
print(f'Headers: {headers_raw}')

col_map = _match_columns(headers_raw)
print(f'col_map: {col_map}')

merge_idx = _find_merge_indices(headers_raw)
print(f'merge_idx: {merge_idx}')
action_idx, target_idx, value_idx = merge_idx

# Debug row 1
row = rows[1]
print(f'\nRow 1: {row}')
print(f'  type: {type(row)}')
print(f'  _cell(row, 1): \"{_cell(row, 1)}\"')
print(f'  _cell(row, 2): \"{_cell(row, 2)}\"')
print(f'  _cell(row, 3): \"{_cell(row, 3)}\"')

# Check action_idx comparsion
values = [str(v).strip() if v is not None else "" for v in row]
print(f'  values: {values}')
print(f'  action_idx({action_idx}) < len(values)({len(values)}): {action_idx < len(values)}')
print(f'  target_idx({target_idx}) < len(values)({len(values)}): {target_idx < len(values)}')
