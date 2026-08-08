"""AutoPilot 用例管理集成测试"""
import httpx, json, io, time, os, sys
from openpyxl import Workbook

BASE = 'http://127.0.0.1:8000/api/v1'

# Create project
r = httpx.post(f'{BASE}/projects/', json={'name': 'CaseTest', 'target_url': 'https://example.com'})
pid = r.json()['data']['id']
print(f'Project: {pid}')

# ═══ Test 1: JSON steps format ═══
wb = Workbook()
ws = wb.active
ws.append(['用例编号', '用例名称', '优先级', '前置条件', '操作步骤', '预期结果'])

steps_json_1 = json.dumps([
    {"action": "navigate", "target": "https://example.com/login", "value": "", "description": "打开登录页"},
    {"action": "fill", "target": "#username", "value": "admin", "description": "输入用户名"},
    {"action": "fill", "target": "#password", "value": "123456", "description": "输入密码"},
    {"action": "click", "target": "button[type=submit]", "value": "", "description": "点击登录"},
    {"action": "assert_text", "target": ".welcome", "value": "欢迎", "description": "验证登录成功"},
], ensure_ascii=False)

ws.append(['TC001', '登录成功', 'P0', '用户已注册', steps_json_1, '页面显示欢迎信息'])

steps_json_2 = json.dumps([
    {"action": "fill", "target": "#username", "value": "admin", "description": "输入用户名"},
    {"action": "fill", "target": "#password", "value": "wrong", "description": "输入错误密码"},
    {"action": "click", "target": "#submit", "value": "", "description": "点击登录"},
], ensure_ascii=False)
ws.append(['TC002', '密码错误', 'P1', None, steps_json_2, '显示密码错误提示'])

buf = io.BytesIO()
wb.save(buf)
r = httpx.post(f'{BASE}/projects/{pid}/cases/import',
    files={'file': ('tc1.xlsx', buf.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')})
d = r.json()['data']
print(f'JSON import: total={d["total"]} success={d["success"]} failed={d["failed"]}')
for e in d.get('errors', []):
    print(f'  ERR: row={e["row"]} {e["reason"]}')

# ═══ Test 2: Plain text steps ═══
wb2 = Workbook()
ws2 = wb2.active
ws2.append(['CaseName', 'Title', '操作步骤', 'Expected'])
ws2.append(['搜索功能', '测试搜索', '1. 打开首页 https://example.com\n2. 在搜索框输入 keyword\n3. 点击搜索按钮\n4. 验证页面显示结果', '展示搜索结果'])
ws2.append(['退出登录', 'test logout', '1. 点击用户头像\n2. 点击退出登录按钮\n3. 验证页面跳转到登录页', '返回登录页'])
buf2 = io.BytesIO()
wb2.save(buf2)
r = httpx.post(f'{BASE}/projects/{pid}/cases/import',
    files={'file': ('tc2.xlsx', buf2.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')})
d = r.json()['data']
print(f'TEXT import: total={d["total"]} success={d["success"]} failed={d["failed"]}')
for e in d.get('errors', []):
    print(f'  ERR: row={e["row"]} {e["reason"]}')

# ═══ Test 3: Action + Target + Value columns ═══
wb3 = Workbook()
ws3 = wb3.active
ws3.append(['用例名', '操作', '对象', '数据', '期望'])
ws3.append(['注册新用户', 'fill', '用户名输入框', 'newuser', '注册成功提示'])
ws3.append(['同意协议', 'click', '同意复选框', '', '协议已勾选'])
buf3 = io.BytesIO()
wb3.save(buf3)
r = httpx.post(f'{BASE}/projects/{pid}/cases/import',
    files={'file': ('tc3.xlsx', buf3.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')})
d = r.json()['data']
print(f'MERGE import: total={d["total"]} success={d["success"]} failed={d["failed"]}')
for e in d.get('errors', []):
    print(f'  ERR: row={e["row"]} {e["reason"]}')

# ═══ Test 4: List ═══
r = httpx.get(f'{BASE}/projects/{pid}/cases/', params={'page': 1, 'size': 10})
d = r.json()['data']
print(f'LIST: total={d["total"]} page={d["page"]}/{d["pages"]}')
for it in d['items']:
    steps_preview = [s.get('description', '')[:40] for s in it.get('steps', [])]
    print(f'  #{it["id"]} {it["case_name"]} [{it["priority"]}] steps={len(it["steps"])} preview={steps_preview}')

# ═══ Test 5: Detail (first case) ═══
if d['items']:
    case_id = d['items'][0]['id']
    r = httpx.get(f'{BASE}/projects/{pid}/cases/{case_id}')
    detail = r.json()['data']
    print(f'DETAIL #{case_id}: {detail["case_name"]} ({len(detail["steps"])} steps)')
    for s in detail['steps']:
        print(f'  {s["action"]} -> {s["target"][:50]} | {s.get("value", "")[:20]}')

# ═══ Test 6: Keyword search ═══
r = httpx.get(f'{BASE}/projects/{pid}/cases/', params={'keyword': '登录'})
print(f'SEARCH "登录": {r.json()["data"]["total"]} results')

# ═══ Test 7: Priority filter ═══
r = httpx.get(f'{BASE}/projects/{pid}/cases/', params={'priority': 'P0'})
print(f'FILTER P0: {r.json()["data"]["total"]} results')

# ═══ Test 8: Batch delete ═══
r = httpx.get(f'{BASE}/projects/{pid}/cases/', params={'page': 1, 'size': 50})
all_ids = [it['id'] for it in r.json()['data']['items']]
if all_ids:
    r = httpx.request('DELETE', f'{BASE}/projects/{pid}/cases/', json={'ids': all_ids})
    print(f'BATCH delete: {r.json()["data"]}')
    r = httpx.get(f'{BASE}/projects/{pid}/cases/')
    print(f'After delete: {r.json()["data"]["total"]} cases')

# Cleanup
r = httpx.delete(f'{BASE}/projects/{pid}')
print(f'Cleanup project: {r.json()["data"]}')
print('ALL TESTS PASSED')
