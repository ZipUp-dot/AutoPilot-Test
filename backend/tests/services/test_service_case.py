"""测试用例管理服务单元测试"""

import io
import json
import pytest
from openpyxl import Workbook
from app.services.case_service import CaseService
from app.models.test_case import TestCase
from app.models.project import Project
from app.exceptions import NotFoundException


class TestCaseServiceImportExcel:
    def test_import_json_steps(self, db_session, mocker):
        mocker.patch("app.services.case_service._save_upload", return_value="/fake/path.xlsx")
        project = Project(name="P", target_url="https://a.com")
        db_session.add(project)
        db_session.commit()

        wb = Workbook()
        ws = wb.active
        ws.append(["用例编号", "用例名称", "操作步骤", "优先级"])
        steps_json = json.dumps([
            {"action": "navigate", "target": "https://example.com", "value": "", "description": "打开"},
            {"action": "click", "target": "#btn", "value": "", "description": "点击"},
        ], ensure_ascii=False)
        ws.append(["TC001", "JSON导入测试", steps_json, "P0"])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        service = CaseService(db_session)
        result = service.import_excel(project.id, buf.getvalue(), "test.xlsx")

        assert result.total == 1
        assert result.success == 1
        assert result.failed == 0

        cases = db_session.query(TestCase).filter(TestCase.project_id == project.id).all()
        assert len(cases) == 1
        assert cases[0].case_name == "JSON导入测试"
        assert cases[0].case_no == "TC001"
        assert cases[0].priority == "P0"
        steps = json.loads(cases[0].steps)
        assert len(steps) == 2
        assert steps[0]["action"] == "navigate"
        assert steps[0]["step_number"] == 1
        assert steps[1]["action"] == "click"
        assert steps[1]["step_number"] == 2

    def test_import_plain_text_steps(self, db_session, mocker):
        mocker.patch("app.services.case_service._save_upload", return_value="/fake/path.xlsx")
        project = Project(name="P", target_url="https://a.com")
        db_session.add(project)
        db_session.commit()

        wb = Workbook()
        ws = wb.active
        ws.append(["用例名称", "操作步骤", "优先级"])
        ws.append([
            "文本步骤测试",
            "1. 打开登录页面 https://example.com/login\n2. 在用户名输入框输入 admin\n3. 点击登录按钮",
            "P1",
        ])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        service = CaseService(db_session)
        result = service.import_excel(project.id, buf.getvalue(), "test.xlsx")

        assert result.success == 1
        cases = db_session.query(TestCase).filter(TestCase.project_id == project.id).all()
        steps = json.loads(cases[0].steps)
        assert len(steps) == 3
        assert steps[0]["action"] == "navigate"
        assert steps[1]["action"] == "fill"
        assert steps[2]["action"] == "click"

    def test_import_action_target_value_columns(self, db_session, mocker):
        mocker.patch("app.services.case_service._save_upload", return_value="/fake/path.xlsx")
        project = Project(name="P", target_url="https://a.com")
        db_session.add(project)
        db_session.commit()

        wb = Workbook()
        ws = wb.active
        ws.append(["用例名称", "操作", "对象", "数据", "优先级"])
        ws.append(["合并列测试", "click", "#submit-btn", "", "P1"])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        service = CaseService(db_session)
        result = service.import_excel(project.id, buf.getvalue(), "test.xlsx")

        assert result.success == 1
        cases = db_session.query(TestCase).filter(TestCase.project_id == project.id).all()
        steps = json.loads(cases[0].steps)
        assert len(steps) == 1
        assert steps[0]["action"] == "click"
        assert steps[0]["target"] == "#submit-btn"

    def test_import_duplicate_case_no_skipped(self, db_session, mocker):
        mocker.patch("app.services.case_service._save_upload", return_value="/fake/path.xlsx")
        project = Project(name="P", target_url="https://a.com")
        db_session.add(project)
        db_session.commit()

        # First import
        wb = Workbook()
        ws = wb.active
        ws.append(["用例编号", "用例名称", "操作步骤"])
        ws.append(["TC001", "用例1", json.dumps([{"action": "click", "target": "#b"}])])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        service = CaseService(db_session)
        result1 = service.import_excel(project.id, buf.getvalue(), "test1.xlsx")
        assert result1.success == 1

        # Second import with same case_no
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.append(["用例编号", "用例名称", "操作步骤"])
        ws2.append(["TC001", "用例1重复", json.dumps([{"action": "click", "target": "#c"}])])
        buf2 = io.BytesIO()
        wb2.save(buf2)
        buf2.seek(0)

        result2 = service.import_excel(project.id, buf2.getvalue(), "test2.xlsx")
        assert result2.success == 0
        assert result2.failed >= 1
        assert any("已存在" in e["reason"] for e in result2.errors)

        # Verify only one case exists
        cases = db_session.query(TestCase).filter(TestCase.project_id == project.id).all()
        assert len(cases) == 1

    def test_import_empty_file(self, db_session, mocker):
        mocker.patch("app.services.case_service._save_upload", return_value="/fake/path.xlsx")
        project = Project(name="P", target_url="https://a.com")
        db_session.add(project)
        db_session.commit()

        wb = Workbook()
        ws = wb.active
        ws.append(["用例名称", "操作步骤"])  # header only, no data rows
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        service = CaseService(db_session)
        result = service.import_excel(project.id, buf.getvalue(), "empty.xlsx")
        assert result.success == 0
        assert result.failed > 0

    def test_import_non_xlsx_extension(self, db_session, mocker):
        mocker.patch("app.services.case_service._save_upload", return_value="/fake/path.txt")
        project = Project(name="P", target_url="https://a.com")
        db_session.add(project)
        db_session.commit()

        service = CaseService(db_session)
        with pytest.raises(Exception):
            service.import_excel(project.id, b"not an excel file", "test.txt")

    def test_import_large_file_handled(self, db_session, mocker):
        mocker.patch("app.services.case_service._save_upload", return_value="/fake/path.xlsx")
        project = Project(name="P", target_url="https://a.com")
        db_session.add(project)
        db_session.commit()

        wb = Workbook()
        ws = wb.active
        ws.append(["用例编号", "用例名称", "操作步骤"])
        for i in range(100):
            ws.append([f"TC{i:04d}", f"用例{i}", json.dumps([{"action": "click", "target": f"#btn{i}"}])])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        service = CaseService(db_session)
        result = service.import_excel(project.id, buf.getvalue(), "large.xlsx")
        assert result.success == 100

    def test_import_missing_required_columns(self, db_session, mocker):
        mocker.patch("app.services.case_service._save_upload", return_value="/fake/path.xlsx")
        project = Project(name="P", target_url="https://a.com")
        db_session.add(project)
        db_session.commit()

        wb = Workbook()
        ws = wb.active
        ws.append(["列A", "列B"])  # no 用例名称 or 操作步骤
        ws.append(["data1", "data2"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        service = CaseService(db_session)
        result = service.import_excel(project.id, buf.getvalue(), "bad.xlsx")
        assert result.success == 0
        assert result.failed > 0


class TestCaseServiceListPaginated:
    def test_with_status_filter(self, db_session):
        project = Project(name="P", target_url="https://a.com")
        db_session.add(project)
        db_session.commit()

        db_session.add(TestCase(project_id=project.id, case_name="C1", steps="[]", status="imported"))
        db_session.add(TestCase(project_id=project.id, case_name="C2", steps="[]", status="pending"))
        db_session.add(TestCase(project_id=project.id, case_name="C3", steps="[]", status="imported"))
        db_session.commit()

        service = CaseService(db_session)
        result = service.list_paginated(project.id, status="imported")
        assert result.total == 2
        assert all(item.status == "imported" for item in result.items)

    def test_with_priority_filter(self, db_session):
        project = Project(name="P", target_url="https://a.com")
        db_session.add(project)
        db_session.commit()

        db_session.add(TestCase(project_id=project.id, case_name="C1", steps="[]", priority="P0"))
        db_session.add(TestCase(project_id=project.id, case_name="C2", steps="[]", priority="P1"))
        db_session.add(TestCase(project_id=project.id, case_name="C3", steps="[]", priority="P1"))
        db_session.commit()

        service = CaseService(db_session)
        result = service.list_paginated(project.id, priority="P1")
        assert result.total == 2
        assert all(item.priority == "P1" for item in result.items)

    def test_with_keyword_search(self, db_session):
        project = Project(name="P", target_url="https://a.com")
        db_session.add(project)
        db_session.commit()

        db_session.add(TestCase(project_id=project.id, case_name="登录测试", case_no="TC001", steps="[]"))
        db_session.add(TestCase(project_id=project.id, case_name="注册测试", case_no="TC002", steps="[]"))
        db_session.add(TestCase(project_id=project.id, case_name="其他", case_no="TC003", steps="[]"))
        db_session.commit()

        service = CaseService(db_session)
        result = service.list_paginated(project.id, keyword="登录")
        assert result.total == 1
        assert result.items[0].case_name == "登录测试"

    def test_keyword_search_on_case_no(self, db_session):
        project = Project(name="P", target_url="https://a.com")
        db_session.add(project)
        db_session.commit()

        db_session.add(TestCase(project_id=project.id, case_name="A", case_no="TC-LOGIN-001", steps="[]"))
        db_session.add(TestCase(project_id=project.id, case_name="B", case_no="TC-OTHER-002", steps="[]"))
        db_session.commit()

        service = CaseService(db_session)
        result = service.list_paginated(project.id, keyword="LOGIN")
        assert result.total == 1
        assert result.items[0].case_no == "TC-LOGIN-001"

    def test_keyword_search_on_steps(self, db_session):
        project = Project(name="P", target_url="https://a.com")
        db_session.add(project)
        db_session.commit()

        steps_json = json.dumps([{"action": "click", "target": "#login-btn", "value": ""}])
        db_session.add(TestCase(project_id=project.id, case_name="Test", steps=steps_json))
        db_session.commit()

        service = CaseService(db_session)
        result = service.list_paginated(project.id, keyword="login-btn")
        assert result.total == 1

    def test_steps_summary_only_first_3(self, db_session):
        project = Project(name="P", target_url="https://a.com")
        db_session.add(project)
        db_session.commit()

        steps = [
            {"step_number": 1, "action": "navigate", "target": "url1", "value": "", "description": "s1"},
            {"step_number": 2, "action": "click", "target": "b1", "value": "", "description": "s2"},
            {"step_number": 3, "action": "fill", "target": "i1", "value": "v", "description": "s3"},
            {"step_number": 4, "action": "click", "target": "b2", "value": "", "description": "s4"},
            {"step_number": 5, "action": "assert_text", "target": "msg", "value": "", "description": "s5"},
        ]
        db_session.add(TestCase(
            project_id=project.id, case_name="MultiStep", steps=json.dumps(steps),
        ))
        db_session.commit()

        service = CaseService(db_session)
        result = service.list_paginated(project.id)
        assert len(result.items) == 1
        assert len(result.items[0].steps_summary) == 3


class TestCaseServiceGetDetail:
    def test_full_steps_returned_not_truncated(self, db_session):
        project = Project(name="P", target_url="https://a.com")
        db_session.add(project)
        db_session.commit()

        steps = [
            {"step_number": 1, "action": "navigate", "target": "url1", "value": "", "description": "s1"},
            {"step_number": 2, "action": "click", "target": "b1", "value": "", "description": "s2"},
            {"step_number": 3, "action": "fill", "target": "i1", "value": "v", "description": "s3"},
            {"step_number": 4, "action": "click", "target": "b2", "value": "", "description": "s4"},
        ]
        case = TestCase(project_id=project.id, case_name="FullSteps", steps=json.dumps(steps))
        db_session.add(case)
        db_session.commit()

        service = CaseService(db_session)
        detail = service.get_detail(project.id, case.id)
        assert len(detail.steps) == 4
        assert detail.steps[0]["action"] == "navigate"
        assert detail.steps[3]["action"] == "click"

    def test_nonexistent_case_raises_not_found(self, db_session):
        project = Project(name="P", target_url="https://a.com")
        db_session.add(project)
        db_session.commit()

        service = CaseService(db_session)
        with pytest.raises(NotFoundException, match="用例 99999 不存在"):
            service.get_detail(project.id, 99999)


class TestCaseServiceDelete:
    def test_delete_one(self, db_session):
        project = Project(name="P", target_url="https://a.com")
        db_session.add(project)
        db_session.commit()

        case = TestCase(project_id=project.id, case_name="ToDelete", steps="[]")
        db_session.add(case)
        db_session.commit()
        case_id = case.id

        service = CaseService(db_session)
        result = service.delete_one(project.id, case_id)
        assert result == case_id

        assert db_session.query(TestCase).filter(TestCase.id == case_id).first() is None

    def test_delete_one_nonexistent(self, db_session):
        project = Project(name="P", target_url="https://a.com")
        db_session.add(project)
        db_session.commit()

        service = CaseService(db_session)
        with pytest.raises(NotFoundException):
            service.delete_one(project.id, 99999)

    def test_delete_batch(self, db_session):
        project = Project(name="P", target_url="https://a.com")
        db_session.add(project)
        db_session.commit()

        ids = []
        for i in range(5):
            case = TestCase(project_id=project.id, case_name=f"C{i}", steps="[]")
            db_session.add(case)
            db_session.flush()
            ids.append(case.id)
        db_session.commit()

        service = CaseService(db_session)
        deleted = service.delete_batch(project.id, ids[:3])
        assert deleted == 3

        remaining = db_session.query(TestCase).filter(TestCase.project_id == project.id).count()
        assert remaining == 2

    def test_delete_batch_some_ids_missing(self, db_session):
        project = Project(name="P", target_url="https://a.com")
        db_session.add(project)
        db_session.commit()

        case = TestCase(project_id=project.id, case_name="OnlyOne", steps="[]")
        db_session.add(case)
        db_session.commit()

        service = CaseService(db_session)
        deleted = service.delete_batch(project.id, [case.id, 99999, 99998])
        assert deleted == 1

        remaining = db_session.query(TestCase).filter(TestCase.project_id == project.id).count()
        assert remaining == 0